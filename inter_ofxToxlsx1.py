# ==========================================================
# SCRIPT COMPLETO: Extrato OFX/PDF → Excel
# Autor: ChatGPT + Jac (colaboração incremental)
# Data da última atualização: 19/09/2025
#
# Funcionalidades:
#   ✅ Importa múltiplos arquivos OFX/PDF
#   ✅ Evita duplicação (FITID no OFX, Data+NrDoc+MEMO no PDF)
#   ✅ Lançamento manual via formulário Tkinter
#   ✅ Opção de desfazer execução
#   ✅ Aba LOG_PROCESSAMENTO para registrar execuções
#   ✅ Aba BANCOS com saldos consolidados por banco/conta
#   ✅ Impressão no console do nome dos arquivos processados
# ==========================================================
# Requisitos:
#   pip install ofxtools openpyxl pdfplumber
# ==========================================================

from ofxtools.Parser import OFXTree
import openpyxl
import os
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
from datetime import datetime
import logging
import pdfplumber
import re
import uuid

# ==========================================================
# Configuração do LOG externo
# ==========================================================
log_file = "processamento_ofx.log"
logging.basicConfig(
    filename=log_file,
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    encoding="utf-8"
)

def safe_get(obj, attr, default=None):
    """Evita KeyError / AttributeError ao acessar atributos"""
    try:
        return getattr(obj, attr)
    except (KeyError, AttributeError):
        return default

# ==========================================================
# Funções auxiliares PDF
# ==========================================================
def parse_valor(valor_str: str) -> float:
    if not valor_str:
        return 0.0
    valor_str = valor_str.strip()
    tipo = valor_str[-1]  # último caractere (C/D)
    numero = valor_str[:-1].strip().replace(".", "").replace(",", ".")
    try:
        valor = float(numero)
    except ValueError:
        valor = 0.0
    if tipo.upper() == "D":
        return -valor
    return valor

def parse_pdf(caminho_pdf):
    dados = []
    padrao = re.compile(r"(\d{2}/\d{2}/\d{4})\s+(\S+)?\s+(.*?)\s+([\d\.,]+ [CD])\s+([\d\.,]+ [CD])")

    with pdfplumber.open(caminho_pdf) as pdf:
        for pagina in pdf.pages:
            tabela = pagina.extract_table()
            if tabela:
                for linha in tabela[1:]:
                    if len(linha) >= 5:
                        data, nr_doc, historico, valor_str, saldo_str = linha[:5]
                        valor = parse_valor(valor_str)
                        saldo = parse_valor(saldo_str)
                        dados.append([data, nr_doc, historico, valor, saldo])
            else:
                texto = pagina.extract_text()
                if not texto:
                    continue
                for linha in texto.split("\n"):
                    m = padrao.match(linha)
                    if m:
                        data, nr_doc, historico, valor_str, saldo_str = m.groups()
                        valor = parse_valor(valor_str)
                        saldo = parse_valor(saldo_str)
                        dados.append([data, nr_doc or "", historico.strip(), valor, saldo])
    return dados

# ==========================================================
# Função: desfazer execução existente
# ==========================================================
def desfazer_execucao(caminho_excel):
    wb = openpyxl.load_workbook(caminho_excel)
    ws = wb.active
    exec_num = simpledialog.askinteger("Desfazer Execução", "Digite o número da execução para remover:")
    if exec_num is None:
        messagebox.showinfo("Cancelado", "Nenhuma execução foi informada.")
        return
    linhas_removidas = 0
    for row in list(ws.iter_rows(min_row=2)):
        if row[7].value == exec_num:  # Execução está na coluna Execução
            ws.delete_rows(row[0].row, 1)
            linhas_removidas += 1
    if linhas_removidas == 0:
        messagebox.showwarning("Aviso", f"Nenhuma linha encontrada para a execução {exec_num}.")
    else:
        wb.save(caminho_excel)
        messagebox.showinfo("Sucesso", f"Foram removidas {linhas_removidas} linhas da execução {exec_num}.")
        logging.info(f"Execução {exec_num} desfeita. Linhas removidas: {linhas_removidas}")

# ==========================================================
# Função: atualizar aba de bancos
# ==========================================================
def atualizar_bancos(wb, ws):
    if "BANCOS" in wb.sheetnames:
        ws_bancos = wb["BANCOS"]
        ws_bancos.delete_rows(2, ws_bancos.max_row)  # limpa os dados antigos
    else:
        ws_bancos = wb.create_sheet("BANCOS")
        ws_bancos.append(["Banco", "Conta", "Saldo Final (R$)"])

    saldos = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:  # Data preenchida
            banco, conta, saldo = row[5], row[6], row[4]
            chave = (banco, conta)
            saldos[chave] = saldo  # pega sempre o último saldo

    total = 0
    for (banco, conta), saldo in saldos.items():
        ws_bancos.append([banco, conta, saldo])
        total += saldo

    ws_bancos.append(["TOTAL GERAL", "", total])

# ==========================================================
# Função: importar OFX/PDF
# ==========================================================
def importar_ofx_pdf(saida):
    caminhos = filedialog.askopenfilenames(
        title="Selecione um ou mais arquivos OFX ou PDF",
        filetypes=[("Arquivos OFX/PDF", "*.ofx *.pdf"), ("Todos os arquivos", "*.*")]
    )
    if not caminhos:
        messagebox.showerror("Erro", "Nenhum arquivo foi selecionado.")
        return

    if os.path.exists(saida):
        wb = openpyxl.load_workbook(saida)
        ws = wb.active
        execucoes = [int(row[7]) for row in ws.iter_rows(min_row=2, values_only=True) if row and isinstance(row[7], int)]
        ultima_execucao = max(execucoes) if execucoes else 0
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Extrato OFX"
        cabecalho = [
            "Data", "Tipo (Entrada/Saída)", "Descrição", "Valor (R$)", "Saldo acumulado (R$)",
            "Banco ID", "Conta", "Execução", "TRNTYPE", "Nr. Documento", "MEMO", "FITID"
        ]
        ws.append(cabecalho)
        ultima_execucao = 0

    if "LOG_PROCESSAMENTO" in wb.sheetnames:
        ws_log = wb["LOG_PROCESSAMENTO"]
    else:
        ws_log = wb.create_sheet("LOG_PROCESSAMENTO")
        ws_log.append([
            "Data Processamento", "Arquivo", "Banco", "Conta", "Execução",
            "Adicionados", "Ignorados", "Entradas (R$)", "Saídas (R$)", "Saldo Final (R$)", "Status"
        ])

    for caminho_arquivo in caminhos:
        print(f"📂 Processando arquivo: {os.path.basename(caminho_arquivo)}")
        execucao_atual = ultima_execucao + 1
        ultima_execucao = execucao_atual

        processados, ignorados, saldo = 0, 0, 0
        data_processo = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        fitids_existentes = set(r[-1] for r in ws.iter_rows(min_row=2, values_only=True) if r and r[-1])

        if caminho_arquivo.lower().endswith(".ofx"):
            # ---- Modo OFX ----
            tree = OFXTree()
            tree.parse(caminho_arquivo)
            ofx = tree.convert()
            stmt = None
            transacoes = []
            if safe_get(ofx, "bankmsgsrsv1"):
                statements = safe_get(ofx.bankmsgsrsv1, "statements")
                if statements and len(statements) > 0:
                    stmt = statements[0]
                    transacoes = safe_get(stmt, "banktranlist", [])
            if not transacoes and safe_get(ofx, "creditcardmsgsrsv1"):
                ccstatement = safe_get(ofx.creditcardmsgsrsv1, "ccstatement")
                if ccstatement:
                    stmt = ccstatement[0].ccstmtrs
                    transacoes = safe_get(stmt, "banktranlist", [])
            if not transacoes:
                messagebox.showerror("Erro", "Não foi possível localizar lançamentos neste OFX.")
                continue

            banco_id = safe_get(safe_get(stmt, "bankacctfrom"), "bankid", "N/A")
            account_id = safe_get(safe_get(stmt, "bankacctfrom"), "acctid", "N/A")
            account_type = safe_get(safe_get(stmt, "bankacctfrom"), "accttype", "N/A")

            for trn in transacoes:
                fitid = safe_get(trn, "fitid")
                if fitid and fitid in fitids_existentes:
                    ignorados += 1
                    continue
                valor = float(trn.trnamt)
                saldo += valor
                tipo = "Entrada" if valor > 0 else "Saída"
                descricao = trn.name or "Sem descrição"
                trntype = safe_get(trn, "trntype", "N/A")
                memo = safe_get(trn, "memo", "")
                nr_doc = ""
                data = trn.dtposted.strftime("%d/%m/%Y")

                ws.append([data, tipo, descricao, valor, saldo, banco_id, account_id, execucao_atual, trntype, nr_doc, memo, fitid])
                processados += 1
                if fitid:
                    fitids_existentes.add(fitid)

        else:
            # ---- Modo PDF ----
            dados = parse_pdf(caminho_arquivo)
            banco_id, account_id = "CEF", "N/A"

            for data_mov, nr_doc, historico, valor, saldo_mov in dados:
                chave = f"{data_mov}-{nr_doc}-{historico}"
                if chave in fitids_existentes:
                    ignorados += 1
                    continue
                saldo += valor
                tipo = "Entrada" if valor > 0 else "Saída"
                trntype = "CREDIT" if valor > 0 else "DEBIT"
                memo = historico
                fitid = f"MANUAL-{execucao_atual}-{uuid.uuid4().int>>96}"

                ws.append([data_mov, tipo, historico, valor, saldo, banco_id, account_id, execucao_atual, trntype, nr_doc, memo, fitid])
                processados += 1
                fitids_existentes.add(chave)

        # Registrar log
        total_entradas = sum(r[3] for r in ws.iter_rows(min_row=2, values_only=True) if r and r[3] > 0)
        total_saidas = sum(r[3] for r in ws.iter_rows(min_row=2, values_only=True) if r and r[3] < 0)

        ws_log.append([
            data_processo,
            os.path.basename(caminho_arquivo),
            banco_id,
            account_id,
            execucao_atual,
            processados,
            ignorados,
            round(total_entradas, 2),
            round(total_saidas, 2),
            round(saldo, 2),
            "OK"
        ])

    atualizar_bancos(wb, ws)
    wb.save(saida)
    messagebox.showinfo("Processo concluído", f"Arquivos importados com sucesso para:\n{saida}")

# ==========================================================
# Função: lançamento manual via formulário
# ==========================================================
def lancamento_manual(saida):
    if not os.path.exists(saida):
        messagebox.showerror("Erro", "Planilha não encontrada! Importe um arquivo antes.")
        return

    wb = openpyxl.load_workbook(saida)
    ws = wb.active
    execucoes = [int(row[7]) for row in ws.iter_rows(min_row=2, values_only=True) if row and isinstance(row[7], int)]
    execucao_atual = max(execucoes) + 1 if execucoes else 1

    def salvar():
        data = entry_data.get()
        tipo = var_tipo.get()
        descricao = entry_desc.get()
        valor = float(entry_valor.get())
        nr_doc = entry_nr_doc.get()
        memo = entry_memo.get()

        saldo = sum(r[3] for r in ws.iter_rows(min_row=2, values_only=True) if r and isinstance(r[3], (int, float)))
        saldo += valor if tipo == "Entrada" else -valor
        trntype = "CREDIT" if tipo == "Entrada" else "DEBIT"
        fitid = f"MANUAL-{execucao_atual}-{uuid.uuid4().int>>96}"
        banco_id, account_id = "MANUAL", "N/A"

        ws.append([data, tipo, descricao, valor if tipo == "Entrada" else -valor, saldo, banco_id, account_id, execucao_atual, trntype, nr_doc, memo, fitid])
        atualizar_bancos(wb, ws)
        wb.save(saida)
        messagebox.showinfo("Sucesso", "Lançamento manual adicionado!")
        root.destroy()

    root = tk.Tk()
    root.title("Lançamento Manual")

    tk.Label(root, text="Data (dd/mm/aaaa):").grid(row=0, column=0)
    entry_data = tk.Entry(root)
    entry_data.grid(row=0, column=1)

    tk.Label(root, text="Tipo:").grid(row=1, column=0)
    var_tipo = tk.StringVar(value="Entrada")
    tk.Radiobutton(root, text="Entrada", variable=var_tipo, value="Entrada").grid(row=1, column=1, sticky="w")
    tk.Radiobutton(root, text="Saída", variable=var_tipo, value="Saída").grid(row=1, column=2, sticky="w")

    tk.Label(root, text="Descrição:").grid(row=2, column=0)
    entry_desc = tk.Entry(root)
    entry_desc.grid(row=2, column=1)

    tk.Label(root, text="Valor:").grid(row=3, column=0)
    entry_valor = tk.Entry(root)
    entry_valor.grid(row=3, column=1)

    tk.Label(root, text="Nr. Documento:").grid(row=4, column=0)
    entry_nr_doc = tk.Entry(root)
    entry_nr_doc.grid(row=4, column=1)

    tk.Label(root, text="Memo:").grid(row=5, column=0)
    entry_memo = tk.Entry(root)
    entry_memo.grid(row=5, column=1)

    tk.Button(root, text="Salvar", command=salvar).grid(row=6, column=0, columnspan=2)

    root.mainloop()

# ==========================================================
# Menu gráfico principal
# ==========================================================
def menu_grafico():
    root = tk.Tk()
    root.title("Gerenciador de Extratos")

    saida = os.path.join(os.getcwd(), "extrato_ofx.xlsx")

    tk.Button(root, text="📥 Importar OFX/PDF", command=lambda: importar_ofx_pdf(saida), width=40).pack(pady=5)
    tk.Button(root, text="📝 Lançamento Manual", command=lambda: lancamento_manual(saida), width=40).pack(pady=5)
    tk.Button(root, text="↩️ Desfazer Execução", command=lambda: desfazer_execucao(saida), width=40).pack(pady=5)
    tk.Button(root, text="🚪 Sair", command=root.destroy, width=40).pack(pady=5)

    root.mainloop()

# ==========================================================
# Execução
# ==========================================================
if __name__ == "__main__":
    menu_grafico()
