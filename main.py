# ==========================================================
# SCRIPT COMPLETÃO: Processar OFX/PDF e salvar em Excel
# Inclui:
#   - Importação múltipla de arquivos OFX/PDF
#   - Prevenção de duplicação (FITID no OFX, Data+NrDoc+MEMO no PDF)
#   - Lançamento manual via formulário (Entrada/Saída)
#   - Função desfazer execução
#   - Aba LOG_PROCESSAMENTO com rastreio por FITID
#   - Aba BANCOS com saldos consolidados
#   - Menu gráfico inicial
# ==========================================================
# Requisitos:
#   pip install ofxtools openpyxl pdfplumber
# ==========================================================

from ofxtools.Parser import OFXTree
import openpyxl
import os
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
from datetime import datetime
import logging
import pdfplumber
import re
import time

# ==========================================================
# Configuração do LOG externo
# ==========================================================
log_file = "processamento_ofx.log"
logging.basicConfig(
    filename=log_file,
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    encoding="utf-8"
)

# ==========================================================
# Helpers
# ==========================================================
def safe_get(obj, attr, default=None):
    """Evita KeyError / AttributeError ao acessar atributos"""
    try:
        return getattr(obj, attr)
    except (KeyError, AttributeError):
        return default

# ==========================================================
# Funções auxiliares PDF
# ==========================================================
def parse_valor(valor_str: str) -> float:
    if not valor_str:
        return 0.0
    valor_str = valor_str.strip()
    tipo = valor_str[-1]  # último caractere (C/D)
    numero = valor_str[:-1].strip().replace(".", "").replace(",", ".")
    try:
        valor = float(numero)
    except ValueError:
        valor = 0.0
    if tipo.upper() == "D":
        return -valor
    return valor

def parse_pdf(caminho_pdf):
    """Extrai lançamentos de um extrato PDF da CEF"""
    dados = []
    padrao = re.compile(r"(\d{2}/\d{2}/\d{4})\s+(\S+)?\s+(.*?)\s+([\d\.,]+ [CD])\s+([\d\.,]+ [CD])")

    with pdfplumber.open(caminho_pdf) as pdf:
        for pagina in pdf.pages:
            tabela = pagina.extract_table()
            if tabela:
                for linha in tabela[1:]:
                    if len(linha) >= 5:
                        data, nr_doc, historico, valor_str, saldo_str = linha[:5]
                        valor = parse_valor(valor_str)
                        saldo = parse_valor(saldo_str)
                        dados.append([data, nr_doc, historico, valor, saldo])
            else:
                texto = pagina.extract_text()
                if not texto:
                    continue
                for linha in texto.split("\n"):
                    m = padrao.match(linha)
                    if m:
                        data, nr_doc, historico, valor_str, saldo_str = m.groups()
                        valor = parse_valor(valor_str)
                        saldo = parse_valor(saldo_str)
                        dados.append([data, nr_doc or "", historico.strip(), valor, saldo])
    return dados

# ==========================================================
# Função: desfazer execução existente
# ==========================================================
def desfazer_execucao(caminho_excel):
    wb = openpyxl.load_workbook(caminho_excel)
    ws = wb.active
    exec_num = simpledialog.askinteger("Desfazer Execução", "Digite o número da execução para remover:")
    if exec_num is None:
        return
    linhas_removidas = 0
    for row in list(ws.iter_rows(min_row=2)):
        if row[7].value == exec_num:  # Execução está na coluna Execução
            ws.delete_rows(row[0].row, 1)
            linhas_removidas += 1
    if linhas_removidas > 0:
        wb.save(caminho_excel)
        messagebox.showinfo("Sucesso", f"Execução {exec_num} desfeita ({linhas_removidas} linhas).")
        logging.info(f"Execução {exec_num} desfeita. Linhas removidas: {linhas_removidas}")
    else:
        messagebox.showwarning("Aviso", f"Nenhuma linha encontrada para a execução {exec_num}.")

# ==========================================================
# Função: atualizar aba BANCOS
# ==========================================================
def atualizar_aba_bancos(wb, ws):
    bancos = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[5] and row[4] is not None:
            chave = (row[5], row[6])  # Banco ID + Processado em
            bancos[chave] = row[4]

    if "BANCOS" in wb.sheetnames:
        wb.remove(wb["BANCOS"])
    ws_bancos = wb.create_sheet("BANCOS")
    ws_bancos.append(["Banco ID", "Conta/Ref", "Saldo Final"])
    for (banco_id, conta), saldo in bancos.items():
        ws_bancos.append([banco_id, conta, saldo])
    ws_bancos.append(["", "TOTAL", sum(bancos.values())])

# ==========================================================
# Criar/abrir planilha
# ==========================================================
def carregar_planilha(saida):
    if os.path.exists(saida):
        wb = openpyxl.load_workbook(saida)
        ws = wb.active
        execucoes = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                if row and row[7] is not None:
                    execucoes.append(int(row[7]))
            except (ValueError, TypeError):
                continue
        ultima_execucao = max(execucoes) if execucoes else 0
        execucao_atual = ultima_execucao + 1
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Extrato OFX"
        cabecalho = [
            "Data", "Tipo (Entrada/Saída)", "Descrição", "Valor (R$)", "Saldo acumulado (R$)",
            "Banco ID", "Processado em", "Execução", "TRNTYPE", "Nr. Documento", "MEMO", "FITID"
        ]
        ws.append(cabecalho)
        execucao_atual = 1
    return wb, ws, execucao_atual

# ==========================================================
# Função: importar múltiplos arquivos
# ==========================================================
def atualizar():
    arquivos = filedialog.askopenfilenames(
        title="Selecione arquivos OFX ou PDF",
        filetypes=[("Arquivos OFX/PDF", "*.ofx *.pdf")]
    )
    if not arquivos:
        return

    saida = os.path.join(os.path.dirname(arquivos[0]), "extrato_ofx.xlsx")
    wb, ws, execucao_atual = carregar_planilha(saida)

    # Aba de log
    if "LOG_PROCESSAMENTO" in wb.sheetnames:
        ws_log = wb["LOG_PROCESSAMENTO"]
    else:
        ws_log = wb.create_sheet("LOG_PROCESSAMENTO")
        ws_log.append([
            "Data Processamento", "Arquivo", "Banco", "Conta", "Execução",
            "Adicionados", "Ignorados", "Entradas (R$)", "Saídas (R$)", "Saldo Final (R$)", "FITID", "Status"
        ])

    fitids_existentes = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row:
            if row[-1]:  # FITID
                fitids_existentes.add(str(row[-1]))
            elif row[-2]:  # MEMO
                chave = f"{row[0]}-{row[-3]}-{row[-2]}"
                fitids_existentes.add(chave)

    total_processados, total_ignorados = 0, 0
    saldo = 0

    for caminho_arquivo in arquivos:
        processados, ignorados = 0, 0
        data_processo = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        if caminho_arquivo.lower().endswith(".ofx"):
            # ---- OFX ----
            tree = OFXTree()
            tree.parse(caminho_arquivo)
            ofx = tree.convert()
            stmt, transacoes = None, []
            if safe_get(ofx, "bankmsgsrsv1"):
                statements = safe_get(ofx.bankmsgsrsv1, "statements")
                if statements and len(statements) > 0:
                    stmt = statements[0]
                    transacoes = safe_get(stmt, "banktranlist", [])
            if not transacoes and safe_get(ofx, "creditcardmsgsrsv1"):
                ccstatement = safe_get(ofx.creditcardmsgsrsv1, "ccstatement")
                if ccstatement:
                    stmt = ccstatement[0].ccstmtrs
                    transacoes = safe_get(stmt, "banktranlist", [])

            banco_id = safe_get(safe_get(stmt, "bankacctfrom"), "bankid", "N/A")
            account_id = safe_get(safe_get(stmt, "bankacctfrom"), "acctid", "N/A")
            account_type = safe_get(safe_get(stmt, "bankacctfrom"), "accttype", "N/A")

            for trn in transacoes:
                fitid = safe_get(trn, "fitid")
                if fitid and fitid in fitids_existentes:
                    ignorados += 1
                    continue
                valor = float(trn.trnamt)
                saldo += valor
                tipo = "Entrada" if valor > 0 else "Saída"
                descricao = trn.name or "Sem descrição"
                trntype = safe_get(trn, "trntype", "N/A")
                memo = safe_get(trn, "memo", "")
                nr_doc = ""
                data = trn.dtposted.strftime("%d/%m/%Y")

                ws.append([data, tipo, descricao, valor, saldo, banco_id, data_processo,
                           execucao_atual, trntype, nr_doc, memo, fitid])
                processados += 1
                if fitid:
                    fitids_existentes.add(fitid)

        else:
            # ---- PDF ----
            dados = parse_pdf(caminho_arquivo)
            banco_id, account_id, account_type = "CEF", "N/A", "N/A"
            for data_mov, nr_doc, historico, valor, saldo_mov in dados:
                chave = f"{data_mov}-{nr_doc}-{historico}"
                if chave in fitids_existentes:
                    ignorados += 1
                    continue
                saldo += valor
                tipo = "Entrada" if valor > 0 else "Saída"
                trntype = "CREDIT" if valor > 0 else "DEBIT"
                memo = historico
                fitid = f"PDF-{execucao_atual}-{int(time.time())}"

                ws.append([data_mov, tipo, historico, valor, saldo, banco_id, data_processo,
                           execucao_atual, trntype, nr_doc, memo, fitid])
                processados += 1
                fitids_existentes.add(chave)

        # Registrar log por arquivo
        ws_log.append([data_processo, os.path.basename(caminho_arquivo),
                       banco_id, account_id, execucao_atual,
                       processados, ignorados,
                       "", "", saldo, "", "OK"])

        total_processados += processados
        total_ignorados += ignorados

    atualizar_aba_bancos(wb, ws)
    wb.save(saida)

    messagebox.showinfo("Processo concluído",
                        f"Arquivos processados: {len(arquivos)}\n"
                        f"Adicionados: {total_processados}\n"
                        f"Ignorados: {total_ignorados}\n"
                        f"Saldo final: {saldo:.2f}")

# ==========================================================
# Função: lançamento manual
# ==========================================================
def lancar_manual(saida):
    wb, ws, execucao_atual = carregar_planilha(saida)

    if "LOG_PROCESSAMENTO" in wb.sheetnames:
        ws_log = wb["LOG_PROCESSAMENTO"]
    else:
        ws_log = wb.create_sheet("LOG_PROCESSAMENTO")
        ws_log.append([
            "Data Processamento", "Arquivo", "Banco", "Conta", "Execução",
            "Adicionados", "Ignorados", "Entradas (R$)", "Saídas (R$)", "Saldo Final (R$)", "FITID", "Status"
        ])

    top = tk.Toplevel()
    top.title("Lançamento Manual")

    tk.Label(top, text="Data (dd/mm/aaaa):").pack()
    entry_data = tk.Entry(top)
    entry_data.pack()
    entry_data.insert(0, datetime.now().strftime("%d/%m/%Y"))

    var_tipo = tk.StringVar(value="Entrada")
    tk.Radiobutton(top, text="Entrada", variable=var_tipo, value="Entrada").pack()
    tk.Radiobutton(top, text="Saída", variable=var_tipo, value="Saída").pack()

    tk.Label(top, text="Descrição:").pack()
    entry_desc = tk.Entry(top)
    entry_desc.pack()

    tk.Label(top, text="Valor:").pack()
    entry_valor = tk.Entry(top)
    entry_valor.pack()

    tk.Label(top, text="Banco ID:").pack()
    entry_banco = tk.Entry(top)
    entry_banco.pack()

    tk.Label(top, text="Nr. Documento:").pack()
    entry_nr_doc = tk.Entry(top)
    entry_nr_doc.pack()

    tk.Label(top, text="MEMO:").pack()
    entry_memo = tk.Entry(top)
    entry_memo.pack()

    def salvar():
        nonlocal execucao_atual
        data_mov = entry_data.get()
        tipo = var_tipo.get()
        descricao = entry_desc.get()
        valor = float(entry_valor.get())
        banco_id = entry_banco.get()
        nr_doc = entry_nr_doc.get()
        memo = entry_memo.get()
        data_processo = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        fitid = f"MANUAL-{execucao_atual}-{int(time.time())}"

        saldo = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[4]:
                saldo = row[4]
        saldo += valor if tipo == "Entrada" else -valor
        trntype = "CREDIT" if tipo == "Entrada" else "DEBIT"

        ws.append([data_mov, tipo, descricao,
                   valor if tipo == "Entrada" else -valor,
                   saldo, banco_id, data_processo, execucao_atual,
                   trntype, nr_doc, memo, fitid])

        ws_log.append([
            data_processo, "MANUAL", banco_id, "N/A",
            execucao_atual, 1, 0,
            valor if tipo == "Entrada" else 0,
            valor if tipo == "Saída" else 0,
            saldo, fitid, "OK"
        ])

        atualizar_aba_bancos(wb, ws)
        wb.save(saida)
        messagebox.showinfo("Sucesso", "Lançamento manual adicionado com sucesso!")
        top.destroy()

    tk.Button(top, text="Salvar", command=salvar).pack()
    top.mainloop()

# ==========================================================
# Menu gráfico
# ==========================================================
def menu_grafico(saida):
    root = tk.Tk()
    root.title("Menu - Processamento Extratos")

    tk.Button(root, text="📥 Atualizar (importar OFX/PDF)", command=atualizar, width=40).pack(pady=5)
    tk.Button(root, text="✍️ Lançamento Manual", command=lambda: lancar_manual(saida), width=40).pack(pady=5)
    tk.Button(root, text="⏪ Desfazer Execução", command=lambda: desfazer_execucao(saida), width=40).pack(pady=5)
    tk.Button(root, text="❌ Sair", command=root.destroy, width=40).pack(pady=5)

    root.mainloop()

# ==========================================================
# Execução principal
# ==========================================================
saida = os.path.join(os.getcwd(), "extrato_ofx.xlsx")
menu_grafico(saida)
