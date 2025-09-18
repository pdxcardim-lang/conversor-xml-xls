# ==========================================================
# SCRIPT: Ler arquivo OFX com ofxtools e salvar em Excel
# Inclui:
#   - Prevenção de duplicação (FITID)
#   - Colunas extras: TRNTYPE e MEMO
#   - Log em arquivo externo + aba LOG_PROCESSAMENTO
#   - Função desfazer execução
# ==========================================================
# Requisitos:
#   pip install ofxtools openpyxl
# ==========================================================

from ofxtools.Parser import OFXTree
import openpyxl
import os
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
from datetime import datetime
import logging

# ==========================================================
# Configuração do LOG externo
# ==========================================================
log_file = "processamento_ofx.log"
logging.basicConfig(
    filename=log_file,
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    encoding="utf-8"
)

def safe_get(obj, attr, default=None):
    """Evita KeyError / AttributeError ao acessar atributos"""
    try:
        return getattr(obj, attr)
    except (KeyError, AttributeError):
        return default

# ==========================================================
# Função: desfazer execução existente
# ==========================================================
def desfazer_execucao(caminho_excel):
    wb = openpyxl.load_workbook(caminho_excel)
    ws = wb.active

    exec_num = simpledialog.askinteger("Desfazer Execução", "Digite o número da execução para remover:")
    if exec_num is None:
        messagebox.showinfo("Cancelado", "Nenhuma execução foi informada.")
        return

    linhas_removidas = 0
    for row in list(ws.iter_rows(min_row=2)):
        if row[-3].value == exec_num:  # Execução está na antepenúltima coluna
            ws.delete_rows(row[0].row, 1)
            linhas_removidas += 1

    if linhas_removidas == 0:
        messagebox.showwarning("Aviso", f"Nenhuma linha encontrada para a execução {exec_num}.")
    else:
        wb.save(caminho_excel)
        messagebox.showinfo("Sucesso", f"Foram removidas {linhas_removidas} linhas da execução {exec_num}.")
        logging.info(f"Execução {exec_num} desfeita. Linhas removidas: {linhas_removidas}")

# ==========================================================
# Seleção do arquivo OFX
# ==========================================================
root = tk.Tk()
root.withdraw()
caminho_arquivo = filedialog.askopenfilename(
    title="Selecione o arquivo OFX",
    filetypes=[("Arquivos OFX", "*.ofx"), ("Todos os arquivos", "*.*")]
)

if not caminho_arquivo:
    messagebox.showerror("Erro", "Nenhum arquivo foi selecionado.")
    logging.error("Nenhum arquivo OFX selecionado.")
    exit()

saida = os.path.join(os.path.dirname(caminho_arquivo), "extrato_ofx.xlsx")

# ==========================================================
# Perguntar se o usuário deseja desfazer execução
# ==========================================================
if os.path.exists(saida):
    opcao = messagebox.askyesno("Opção", "Deseja desfazer uma execução existente?\n(Sim = desfazer, Não = importar novo OFX)")
    if opcao:
        desfazer_execucao(saida)
        exit()

# ==========================================================
# Ler OFX
# ==========================================================
try:
    tree = OFXTree()
    tree.parse(caminho_arquivo)
    ofx = tree.convert()
except Exception as e:
    messagebox.showerror("Erro", f"Falha ao ler o OFX: {e}")
    logging.error(f"Erro ao ler OFX {caminho_arquivo}: {e}")
    exit()

# ==========================================================
# Obter transações (Nubank usa statements[0].banktranlist)
# ==========================================================
transacoes = []
stmt = None

if safe_get(ofx, "bankmsgsrsv1"):
    statements = safe_get(ofx.bankmsgsrsv1, "statements")
    if statements and len(statements) > 0:
        stmt = statements[0]
        transacoes = safe_get(stmt, "banktranlist", [])

# Fallback cartão
if not transacoes and safe_get(ofx, "creditcardmsgsrsv1"):
    ccstatement = safe_get(ofx.creditcardmsgsrsv1, "ccstatement")
    if ccstatement:
        stmt = ccstatement[0].ccstmtrs
        transacoes = safe_get(stmt, "banktranlist", [])

# ==========================================================
# Validar se encontrou transações
# ==========================================================
if not transacoes:
    messagebox.showerror("Erro", "Não foi possível localizar lançamentos neste OFX.")
    logging.error(f"OFX {caminho_arquivo}: Nenhum lançamento encontrado.")
    exit()

# ==========================================================
# Info da conta/banco
# ==========================================================
banco_id = safe_get(safe_get(stmt, "bankacctfrom"), "bankid", "N/A")
account_id = safe_get(safe_get(stmt, "bankacctfrom"), "acctid", "N/A")
account_type = safe_get(safe_get(stmt, "bankacctfrom"), "accttype", "N/A")

# ==========================================================
# Criar/abrir planilha
# ==========================================================
if os.path.exists(saida):
    wb = openpyxl.load_workbook(saida)
    ws = wb.active
    execucoes = [int(row[7]) for row in ws.iter_rows(min_row=2, values_only=True)
    if row and row[7] is not None]
    ultima_execucao = max(execucoes) if execucoes else 0
    execucao_atual = ultima_execucao + 1
else:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extrato OFX"
    cabecalho = [
        "Data", "Tipo (Entrada/Saída)", "Descrição", "Valor (R$)", "Saldo acumulado (R$)",
        "Banco ID", "Processado em", "Execução", "TRNTYPE", "MEMO", "FITID"
    ]
    ws.append(cabecalho)
    execucao_atual = 1

# Criar/abrir aba de log
if "LOG_PROCESSAMENTO" in wb.sheetnames:
    ws_log = wb["LOG_PROCESSAMENTO"]
else:
    ws_log = wb.create_sheet("LOG_PROCESSAMENTO")
    ws_log.append([
        "Data Processamento", "Arquivo OFX", "Banco", "Conta", "Execução",
        "Adicionados", "Ignorados", "Entradas (R$)", "Saídas (R$)", "Saldo Final (R$)", "Status"
    ])

# ==========================================================
# Evitar duplicação
# ==========================================================
fitids_existentes = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    if row and row[-1]:
        fitids_existentes.add(str(row[-1]))

# ==========================================================
# Processar transações
# ==========================================================
saldo = 0
processados = 0
ignorados = 0
data_processo = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

for trn in transacoes:
    fitid = safe_get(trn, "fitid")

    if fitid and fitid in fitids_existentes:
        ignorados += 1
        continue

    valor = float(trn.trnamt)
    saldo += valor
    tipo = "Entrada" if valor > 0 else "Saída"
    descricao = trn.name if trn.name else "Sem descrição"
    trntype = safe_get(trn, "trntype", "N/A")
    memo = safe_get(trn, "memo", "")

    data = trn.dtposted.strftime("%d/%m/%Y")

    ws.append([
        data,
        tipo,
        descricao,
        valor,
        saldo,
        banco_id,
        data_processo,
        execucao_atual,
        trntype,
        memo,
        fitid
    ])
    processados += 1
    if fitid:
        fitids_existentes.add(fitid)

# ==========================================================
# Totais
# ==========================================================
total_entradas = sum(float(t.trnamt) for t in transacoes if float(t.trnamt) > 0)
total_saidas = sum(float(t.trnamt) for t in transacoes if float(t.trnamt) < 0)
saldo_final = saldo

# ==========================================================
# Registrar log
# ==========================================================
ws_log.append([
    data_processo,
    os.path.basename(caminho_arquivo),
    banco_id,
    account_id,
    execucao_atual,
    processados,
    ignorados,
    round(total_entradas, 2),
    round(total_saidas, 2),
    round(saldo_final, 2),
    "OK"
])

wb.save(saida)

mensagem = (
    f"Planilha atualizada com sucesso:\n{saida}\n\n"
    f"Banco: {banco_id} | Conta: {account_id} ({account_type})\n"
    f"Transações adicionadas: {processados}\n"
    f"Transações ignoradas (duplicadas): {ignorados}\n"
    f"Execução nº {execucao_atual}\n"
    f"Resumo:\n"
    f" - Saldo final: R$ {saldo_final:.2f}\n"
    f" - Entradas: R$ {total_entradas:.2f}\n"
    f" - Saídas: R$ {total_saidas:.2f}"
)

print(mensagem)
messagebox.showinfo("Processo concluído", mensagem)

logging.info(
    f"Arquivo OFX: {caminho_arquivo} | Banco: {banco_id} | Conta: {account_id} "
    f"| Execução: {execucao_atual} | Adicionados: {processados} | Ignorados: {ignorados} "
    f"| Entradas: {total_entradas:.2f} | Saídas: {total_saidas:.2f} | Saldo final: {saldo_final:.2f}"
)
